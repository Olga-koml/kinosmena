# работает через pandas но не могу добавить название в начале перед таблицей
# import pandas as pd
# from django.http import HttpResponse
# from .models import Shift
# from users.models import TelegramUser
# from io import BytesIO
# from openpyxl.utils.dataframe import dataframe_to_rows

# def get_export_shifts_to_excel(request):
#     tid = request.GET.get('tid')
#     user = TelegramUser.objects.get(tid=tid)
#     shifts = Shift.objects.filter(user=user)


#     for shift in shifts:
#         shift.start_date = shift.start_date.astimezone(pytz.utc).replace(tzinfo=None)
#         shift.end_date = shift.end_date.astimezone(pytz.utc).replace(tzinfo=None)


#     # Создаем DataFrame из данных
#     data = {

#         'Название проекта': [shift.project.name for shift in shifts],
#         'Начало смены': [shift.start_date.strftime('%d.%m.%Y %H:%M') for shift in shifts],
#         'Окончание смены': [shift.end_date.strftime('%d.%m.%Y %H:%M') for shift in shifts],
#         # 'Начало смены': [shift.start_date for shift in shifts],
#         # 'End Date': [shift.end_date for shift in shifts],
#         'Стоимость смены': [shift.shift_sum for shift in shifts],
#         # 'Сумма переработок': [shift.overwork_sum for shift in shifts],
#         Shift._meta.get_field('overwork_sum').verbose_name: [shift.overwork_sum for shift in shifts],
#         Shift._meta.get_field('overwork_hours').verbose_name: [shift.overwork_hours for shift in shifts],
#         Shift._meta.get_field('non_sleep_sum').verbose_name: [shift.non_sleep_sum for shift in shifts],
#         Shift._meta.get_field('non_sleep_hours').verbose_name: [shift.non_sleep_hours for shift in shifts],
#         Shift._meta.get_field('day_off_sum').verbose_name: [shift.day_off_sum for shift in shifts],
#         Shift._meta.get_field('day_off_hours').verbose_name: [shift.day_off_hours for shift in shifts],
#         Shift._meta.get_field('is_day_off').verbose_name: [1 if shift.is_day_off else 0 for shift in shifts],
#         Shift._meta.get_field('current_lunch_sum').verbose_name: [shift.current_lunch_sum for shift in shifts],
#         Shift._meta.get_field('is_current_lunch').verbose_name: [1 if shift.is_current_lunch else 0 for shift in shifts],
#         Shift._meta.get_field('late_lunch_sum').verbose_name: [shift.late_lunch_sum for shift in shifts],
#         Shift._meta.get_field('is_late_lunch').verbose_name: [1 if shift.is_late_lunch else 0 for shift in shifts],
#         Shift._meta.get_field('per_diem_sum').verbose_name: [shift.per_diem_sum for shift in shifts],
#         Shift._meta.get_field('is_per_diem').verbose_name: [1 if shift.is_per_diem else 0 for shift in shifts],
#         Shift._meta.get_field('services_sum').verbose_name: [shift.services_sum for shift in shifts],
#         Shift._meta.get_field('total').verbose_name: [shift.total for shift in shifts],
#         'Сумма с учетом 6%': [shift.total*1.06 for shift in shifts]
#     }

#     # df = pd.DataFrame(data)

#     # # Создаем Excel-файл в памяти
#     # output = BytesIO()
#     # writer = pd.ExcelWriter(output, engine='xlsxwriter')
#     # df.to_excel(writer, index=True, sheet_name='Shifts')
#     # writer.close()

#     # response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     # response['Content-Disposition'] = 'attachment; filename=shifts.xlsx'

#     # return response


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from .models import Shift
from users.models import TelegramUser
from io import BytesIO
import pytz
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.conf import settings
from bot import send_message_async, send_telegram_message
from asgiref.sync import async_to_sync
import math


def get_format_date(date):
    """Преобразует дату в формат ДД.MM.ГГГГ Ч:М."""
    return date.astimezone(pytz.utc).replace(tzinfo=None).strftime('%d.%m.%Y %H:%M')


def get_export_shifts_to_text_report(object):
    message = ''
    total_hours = 0
  
    tid = object.user

    # start_date = object.start_date.astimezone(pytz.utc).replace(tzinfo=None).strftime('%d.%m.%Y %H:%M')
    # end_date = object.end_date.astimezone(pytz.utc).replace(tzinfo=None).strftime('%d.%m.%Y %H:%M')
    start_date = get_format_date(object.start_date)
    end_date = get_format_date(object.end_date)
    fact_shift_duration = math.ceil(
            ((object.end_date - object.start_date).total_seconds() / 3600)
        )
    message += (
        f'Проект: {object.project.name}\n'
        f'Начало смены: {start_date}\n'
        f'Конец смены: {end_date}\n'
        f'Фактическая продолжительность смены: {fact_shift_duration} ч\n'
    )
    if object.overwork_hours > 0:
        message += f'Переработки: {object.overwork_hours} ч\n'
        total_hours += object.overwork_hours
    if object.non_sleep_hours > 0:
        message += f'Недосып: {object.non_sleep_hours} ч\n'
        total_hours += object.non_sleep_hours
    if object.day_off_hours > 0:
        message += f'Day_off: {object.day_off_hours} ч\n'
        total_hours += object.day_off_hours
    if object.is_current_lunch:
        message += 'Текущий обед: 1 ч\n'
        total_hours += 1
    if object.is_late_lunch:
        message += 'Поздний обед: 1 ч\n'
        total_hours += 1
    if object.is_per_diem:
        message += 'Суточные: да\n'
    if total_hours > 0:
        message += f'Итого сверх плановой смены: {total_hours} ч'
    print(message)
    # async_send_message = async_to_sync(send_message_async)  # Оборачиваем асинхронную функцию в синхронную
    # async_send_message(message=message, telegram_id=tid)
    send_telegram_message(chat_id=tid, message=message)
    # await send_message_async(message=message, telegram_id=tid)
    response = HttpResponse(
        message,
        content_type='text/plain'
        # content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response


def get_export_shifts_to_excel(queryset):
    # tid = request.GET.get('tid')
    # user = TelegramUser.objects.get(tid=tid)
    # shifts = Shift.objects.filter(user=user).order_by('start_date')
    shifts = queryset.order_by('start_date')
    
    for shift in shifts:
        # shift.start_date = shift.start_date.astimezone(pytz.utc).replace(tzinfo=None)
        # shift.end_date = shift.end_date.astimezone(pytz.utc).replace(tzinfo=None)
        shift.start_date = get_format_date(shift.start_date)
        shift.end_date = get_format_date(shift.end_date)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Добавляем заголовки перед таблицей
    ws.append(['Kinosmena'])
    ws.append(['Отчет по всем сменам'])
    ws.append([])  # Пустая строка

    # Добавляем заголовки столбцов
    ws.append([
        'Номер',
        'Название проекта',
        'Кол-во смен',
        'Начало смены',
        'Окончание смены',
        'Стоимость смены',
        'Cтоимость переработки в час',
        Shift._meta.get_field('overwork_hours').verbose_name,
        Shift._meta.get_field('overwork_sum').verbose_name,
        'Стоимость недосыпа в час',
        Shift._meta.get_field('non_sleep_hours').verbose_name,
        Shift._meta.get_field('non_sleep_sum').verbose_name,
        'Cтоимость day_off в час',
        Shift._meta.get_field('is_day_off').verbose_name + ', 1 - да, 0 - нет',
        Shift._meta.get_field('day_off_hours').verbose_name,
        Shift._meta.get_field('day_off_sum').verbose_name,
        Shift._meta.get_field('is_current_lunch').verbose_name + ', 1 - да, 0 - нет',
        Shift._meta.get_field('current_lunch_sum').verbose_name,
        Shift._meta.get_field('is_late_lunch').verbose_name + ', 1 - да, 0 - нет',
        Shift._meta.get_field('late_lunch_sum').verbose_name,
        Shift._meta.get_field('is_per_diem').verbose_name + ', 1 - да, 0 - нет',
        Shift._meta.get_field('per_diem_sum').verbose_name,
        Shift._meta.get_field('services_sum').verbose_name,
        Shift._meta.get_field('total').verbose_name,
        # 'Сумма с учетом 6%'
    ])

    # Устанавливаем стили для заголовков
    header_font = Font(bold=True)
    for cell in ws[4]:
        cell.font = header_font
    # Создаем объект заливки
    fill = PatternFill(start_color='F6F298', end_color='F6F298', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    # Применяем заливку к каждой ячейке в строке с заголовками (например, с 4-ой строки)
    for row in ws.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.fill = fill
            cell.alignment = alignment
    # Добавляем данные
    for idx, shift in enumerate(shifts, start=1):
        ws.append([
            idx,
            shift.project.name,
            1,
            # shift.start_date.strftime('%d.%m.%Y %H:%M'),
            # shift.end_date.strftime('%d.%m.%Y %H:%M'),
            shift.start_date,
            shift.end_date,
            shift.shift_sum,
            shift.project.overtime_rate,
            shift.overwork_hours,
            shift.overwork_sum,
            shift.project.non_sleep_rate,
            shift.non_sleep_hours,
            shift.non_sleep_sum,
            shift.project.day_off_rate,
            1 if shift.is_day_off else 0,
            shift.day_off_hours,
            shift.day_off_sum,
            1 if shift.is_current_lunch else 0,
            shift.current_lunch_sum,
            1 if shift.is_late_lunch else 0,
            shift.late_lunch_sum,
            1 if shift.is_per_diem else 0,
            shift.per_diem_sum,
            shift.services_sum,
            shift.total,
            # shift.total*1.06
        ])

    # Итоговая строка считается через ексель
    current_max_row = ws.max_row
    for col in range(3, ws.max_column+1):
        col_letter = get_column_letter(col)
        sum_formula = f'=SUM({col_letter}5:{col_letter}{current_max_row})'
        ws[f'{col_letter}{current_max_row + 1}'] = sum_formula

    ws[f'B{current_max_row + 1}'] = 'ИТОГО'
    # Устанавливаем стили для ячеек с данными
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True, vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # всем столбцам устанавливаю ширину 18
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 18

    # делаем ширину первого столбца поменьше
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['C'].width = 8

    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill
            cell.font = header_font

    # Создаем файл в памяти
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=shifts.xlsx'

    return response
